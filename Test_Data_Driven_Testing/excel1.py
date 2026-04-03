import pytest
from openpyxl.workbook import Workbook
import openpyxl
wb=openpyxl.Workbook()
sheetName="sheet1"
if sheetName in wb.sheetnames:
    ws=wb[sheetName]
else:
    ws=wb.create_sheet(sheetName)

ws['A1']='user'
ws['B1']='password'

wb.save('sample.xlsx')  # IT WILL BE SAVED IN LOCAL DIRECTORY
ws.append(['user1','123'])
ws.append(['user2','1232'])
ws.append(['user3','1261'])
ws.append(['user4','1233'])
ws.append(['standard_user','secret_sauce'])

wb.save('sample.xlsx')

#to print and to iterate the data of a excel file
for row in ws.iter_rows(values_only=True):
    print(row)
